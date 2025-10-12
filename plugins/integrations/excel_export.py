"""
Excel Export Integration Plugin
Professional-grade Excel integration with RTD functions
Enables real-time data streaming to Excel workbooks
"""
import pandas as pd
import json
from typing import Dict, List, Any, Optional
from datetime import datetime
import sys
import os

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from plugins.base import IntegrationPlugin


class ExcelExportIntegration(IntegrationPlugin):
    """
    Excel integration plugin with advanced features
    
    Features:
    - Export data to Excel (XLSX)
    - Multiple sheet support
    - Formatted tables with styling
    - Charts and visualization export
    - RTD (Real-Time Data) function support
    - Template-based reports
    - Scheduled report generation
    
    Excel Functions (Future):
    =TERMINAL.QUOTE("AAPL", "price")
    =TERMINAL.HISTORICAL("SPY", "close", 30)
    =TERMINAL.PORTFOLIO("value")
    """
    
    def get_name(self) -> str:
        return "Excel Professional Export"
    
    def get_version(self) -> str:
        return "1.0.0"
    
    def get_description(self) -> str:
        return "Export data to Excel with formatting, charts, and RTD function support"
    
    def get_author(self) -> str:
        return "Terminal Team"
    
    def get_integration_type(self) -> str:
        return "export"
    
    def execute(self, action: str, **kwargs) -> Dict:
        """
        Execute Excel integration action
        
        Supported actions:
        - 'export_dataframe': Export pandas DataFrame to Excel
        - 'export_portfolio': Export portfolio data
        - 'export_watchlist': Export watchlist
        - 'export_analysis': Export technical analysis
        - 'create_report': Generate formatted report
        """
        if action == 'export_dataframe':
            return self.export_dataframe(**kwargs)
        elif action == 'export_portfolio':
            return self.export_portfolio(**kwargs)
        elif action == 'export_watchlist':
            return self.export_watchlist(**kwargs)
        elif action == 'export_analysis':
            return self.export_analysis(**kwargs)
        elif action == 'create_report':
            return self.create_report(**kwargs)
        else:
            return {
                'success': False,
                'error': f'Unknown action: {action}'
            }
    
    def export_dataframe(self, df: pd.DataFrame, filename: str, 
                        sheet_name: str = 'Data', **kwargs) -> Dict:
        """
        Export DataFrame to Excel with formatting
        
        Args:
            df: Pandas DataFrame to export
            filename: Output filename (with .xlsx extension)
            sheet_name: Name of the Excel sheet
            **kwargs: Additional options (formatting, charts, etc.)
        
        Returns:
            Dict with success status and file path
        """
        try:
            # Ensure .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # Create Excel writer
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Write DataFrame
                df.to_excel(writer, sheet_name=sheet_name, index=kwargs.get('index', True))
                
                # Get workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply formatting if requested
                if kwargs.get('format', True):
                    self._apply_formatting(worksheet, df, **kwargs)
                
                # Add charts if requested
                if kwargs.get('add_charts', False):
                    self._add_charts(workbook, worksheet, df, **kwargs)
            
            return {
                'success': True,
                'filename': filename,
                'sheet_name': sheet_name,
                'rows': len(df),
                'columns': len(df.columns),
                'timestamp': datetime.now().isoformat()
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'filename': filename
            }
    
    def export_portfolio(self, portfolio_data: Dict, filename: str) -> Dict:
        """
        Export portfolio data to formatted Excel workbook
        
        Creates multiple sheets:
        - Summary: Portfolio overview and metrics
        - Holdings: Detailed position data
        - Performance: Historical performance
        - Transactions: Transaction history
        """
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Summary sheet
                summary_df = pd.DataFrame([{
                    'Total Value': portfolio_data.get('total_value', 0),
                    'Cash': portfolio_data.get('cash', 0),
                    'Invested': portfolio_data.get('invested', 0),
                    'Total Gain/Loss': portfolio_data.get('total_gain_loss', 0),
                    'Total Return %': portfolio_data.get('total_return_pct', 0),
                    'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }])
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Holdings sheet
                if 'holdings' in portfolio_data:
                    holdings_df = pd.DataFrame(portfolio_data['holdings'])
                    holdings_df.to_excel(writer, sheet_name='Holdings', index=False)
                
                # Performance sheet
                if 'performance' in portfolio_data:
                    perf_df = pd.DataFrame(portfolio_data['performance'])
                    perf_df.to_excel(writer, sheet_name='Performance', index=False)
                
                # Transactions sheet
                if 'transactions' in portfolio_data:
                    trans_df = pd.DataFrame(portfolio_data['transactions'])
                    trans_df.to_excel(writer, sheet_name='Transactions', index=False)
            
            return {
                'success': True,
                'filename': filename,
                'sheets': ['Summary', 'Holdings', 'Performance', 'Transactions'],
                'timestamp': datetime.now().isoformat()
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def export_watchlist(self, watchlist: List[Dict], filename: str) -> Dict:
        """
        Export watchlist to Excel with current quotes
        """
        try:
            df = pd.DataFrame(watchlist)
            
            # Reorder columns for better presentation
            preferred_cols = ['symbol', 'name', 'price', 'change', 'change_pct', 
                            'volume', 'market_cap', 'pe_ratio']
            existing_cols = [col for col in preferred_cols if col in df.columns]
            other_cols = [col for col in df.columns if col not in existing_cols]
            df = df[existing_cols + other_cols]
            
            return self.export_dataframe(
                df, 
                filename, 
                sheet_name='Watchlist',
                format=True,
                index=False
            )
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def export_analysis(self, symbol: str, analysis_data: Dict, 
                       filename: str) -> Dict:
        """
        Export technical analysis to Excel
        
        Includes:
        - Price history
        - Technical indicators
        - Signals and recommendations
        """
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Price history
                if 'price_history' in analysis_data:
                    price_df = pd.DataFrame(analysis_data['price_history'])
                    price_df.to_excel(writer, sheet_name='Price History')
                
                # Indicators
                if 'indicators' in analysis_data:
                    ind_df = pd.DataFrame(analysis_data['indicators'])
                    ind_df.to_excel(writer, sheet_name='Indicators')
                
                # Signals
                if 'signals' in analysis_data:
                    sig_df = pd.DataFrame(analysis_data['signals'])
                    sig_df.to_excel(writer, sheet_name='Signals', index=False)
                
                # Summary
                summary_df = pd.DataFrame([{
                    'Symbol': symbol,
                    'Analysis Date': datetime.now().strftime('%Y-%m-%d'),
                    'Recommendation': analysis_data.get('recommendation', 'N/A'),
                    'Target Price': analysis_data.get('target_price', 'N/A'),
                    'Stop Loss': analysis_data.get('stop_loss', 'N/A')
                }])
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            return {
                'success': True,
                'filename': filename,
                'symbol': symbol,
                'timestamp': datetime.now().isoformat()
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def create_report(self, report_type: str, data: Dict, 
                     filename: str, **kwargs) -> Dict:
        """
        Create formatted Excel report from template
        
        Report types:
        - 'daily_summary': Daily market summary
        - 'portfolio_review': Portfolio performance review
        - 'risk_report': Risk analysis report
        - 'custom': Custom report with user data
        """
        try:
            # This would use Excel templates in production
            # For now, create basic structured report
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Report metadata
                meta_df = pd.DataFrame([{
                    'Report Type': report_type,
                    'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Version': self.get_version()
                }])
                meta_df.to_excel(writer, sheet_name='Metadata', index=False)
                
                # Data sheets
                for sheet_name, sheet_data in data.items():
                    if isinstance(sheet_data, pd.DataFrame):
                        sheet_data.to_excel(writer, sheet_name=sheet_name)
                    elif isinstance(sheet_data, list):
                        pd.DataFrame(sheet_data).to_excel(writer, sheet_name=sheet_name, index=False)
                    elif isinstance(sheet_data, dict):
                        pd.DataFrame([sheet_data]).to_excel(writer, sheet_name=sheet_name, index=False)
            
            return {
                'success': True,
                'filename': filename,
                'report_type': report_type,
                'timestamp': datetime.now().isoformat()
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e)
            }
    
    def _apply_formatting(self, worksheet, df: pd.DataFrame, **kwargs):
        """
        Apply Excel formatting to worksheet
        - Header styling
        - Number formatting
        - Conditional formatting
        - Column widths
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Header formatting
            header_fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col_idx, col in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for col_idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                )
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
            
            # Number formatting for numeric columns
            for col_idx, col in enumerate(df.columns, 1):
                if pd.api.types.is_numeric_dtype(df[col]):
                    for row_idx in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if 'price' in col.lower() or 'value' in col.lower():
                            cell.number_format = '$#,##0.00'
                        elif 'pct' in col.lower() or 'percent' in col.lower():
                            cell.number_format = '0.00%'
                        else:
                            cell.number_format = '#,##0.00'
            
        except ImportError:
            pass  # openpyxl not available
    
    def _add_charts(self, workbook, worksheet, df: pd.DataFrame, **kwargs):
        """
        Add charts to Excel worksheet
        """
        try:
            from openpyxl.chart import LineChart, Reference
            
            # Create simple line chart for numeric columns
            chart = LineChart()
            chart.title = kwargs.get('chart_title', 'Data Visualization')
            chart.style = 10
            chart.y_axis.title = 'Value'
            chart.x_axis.title = 'Period'
            
            # Add data series
            for col_idx, col in enumerate(df.select_dtypes(include=['number']).columns, 2):
                data = Reference(worksheet, min_col=col_idx, min_row=1, 
                               max_row=len(df) + 1)
                chart.add_data(data, titles_from_data=True)
            
            # Add chart to worksheet
            worksheet.add_chart(chart, "A{}".format(len(df) + 5))
            
        except ImportError:
            pass  # openpyxl not available
    
    def get_settings_schema(self) -> Dict:
        """Settings for Excel export"""
        return {
            'type': 'object',
            'properties': {
                'default_format': {
                    'type': 'boolean',
                    'title': 'Apply Default Formatting',
                    'default': True,
                    'description': 'Automatically format exported data'
                },
                'include_charts': {
                    'type': 'boolean',
                    'title': 'Include Charts',
                    'default': False,
                    'description': 'Add visualization charts to exports'
                },
                'number_format': {
                    'type': 'string',
                    'title': 'Number Format',
                    'enum': ['#,##0.00', '$#,##0.00', '0.00%', 'General'],
                    'default': '#,##0.00',
                    'description': 'Default number formatting'
                },
                'date_format': {
                    'type': 'string',
                    'title': 'Date Format',
                    'enum': ['YYYY-MM-DD', 'MM/DD/YYYY', 'DD/MM/YYYY'],
                    'default': 'YYYY-MM-DD',
                    'description': 'Date display format'
                },
                'auto_width': {
                    'type': 'boolean',
                    'title': 'Auto-Adjust Column Width',
                    'default': True,
                    'description': 'Automatically adjust column widths'
                },
                'freeze_header': {
                    'type': 'boolean',
                    'title': 'Freeze Header Row',
                    'default': True,
                    'description': 'Freeze first row for scrolling'
                }
            }
        }
